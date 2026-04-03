#!/usr/bin/env python3
"""
Second-pass open-data reconstruction for
"Analytical Decomposition of Nodal Frequency Responses".

This version follows the cross-paper clues from the authors' Applied Energy
paper more closely:
- all original IEEE-118 lines use x_new = 0.4 * x_old
- the 54 generator-side nodes are parameterized with the reported
  27xTB6, 9xTC6, 5xTB10, 2xTC10, 11xTG3 fleet mix
- 100 MVA base and 50 Hz conversion are kept explicit
- load damping starts from 1% of load power
- buses 37 / 49 / 50 and the northern weak area around 8 / 9 are given
  focused structural adjustments for an open-data fit

This remains a transparent engineering reconstruction, not an exact recovery of
non-public author data.
"""
from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

import numpy as np
import pandas as pd
import scipy.linalg as la

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from pypower.case118 import case118
except Exception as exc:  # pragma: no cover
    raise SystemExit("Install pypower first: pip install pypower") from exc


@dataclass(frozen=True)
class TypeTemplate:
    name: str
    count: int
    capacity_mw: float
    inertia_s: float
    xdp: float
    deadband_hz: float
    delay_s: float


@dataclass
class ClusterRule:
    hosts: List[int]
    transformer_x_mult: float = 1.0
    inertia_mult: float = 1.0
    damping_mult: float = 1.0
    governor_gain_mult: float = 1.0
    tau_mult: float = 1.0


@dataclass
class ReproConfigV2:
    name: str = "ae_informed_rocof_fit"
    base_mva: float = 100.0
    base_frequency_hz: float = 50.0
    line_x_scale: float = 0.4
    disturbance_load_bus: int = 50
    disturbance_mw: float = 1200.0
    duration_s: float = 20.0
    n_points: int = 2001
    governor_gamma: float = 0.20
    governor_tau_s: float = 0.75
    governor_gain_scale: float = 0.90
    droop_R: float = 0.05
    inertia_scale: float = 1.05
    damping_pu_fraction: float = 0.02
    load_damping_fraction: float = 0.01
    min_load_for_mu_mw: float = 10.0

    # Open-data fit to recover the paper's RoCoF geography more closely.
    north_corridor_scale: float = 0.55
    north_edges: List[Tuple[int, int]] = field(default_factory=lambda: [
        (8, 9), (8, 5), (8, 30), (9, 10), (4, 11), (5, 11),
        (11, 12), (11, 13), (30, 17), (26, 30), (30, 38),
    ])
    clusters: List[ClusterRule] = field(default_factory=lambda: [
        # Northern weak-support area: make 8/9 prone to deeper nadirs but keep
        # their initial RoCoF small because they are far from the disturbance.
        ClusterRule(hosts=[8, 10, 12], transformer_x_mult=1.25,
                    inertia_mult=1.60, damping_mult=0.90,
                    governor_gain_mult=0.42, tau_mult=2.00),
        # Disturbance corridor: make the bus-49 neighbourhood show the largest
        # initial RoCoF while preserving reasonable post-fault support.
        ClusterRule(hosts=[46], transformer_x_mult=0.95, inertia_mult=0.85,
                    damping_mult=1.00, governor_gain_mult=1.40, tau_mult=0.75),
        ClusterRule(hosts=[49], transformer_x_mult=0.72, inertia_mult=0.24,
                    damping_mult=1.10, governor_gain_mult=3.00, tau_mult=0.40),
        # Strong support south-east of the disturbance so 54-56 do not dominate
        # the RoCoF ranking even though they are electrically close.
        ClusterRule(hosts=[54, 55, 56], transformer_x_mult=0.85,
                    inertia_mult=2.80, damping_mult=1.20,
                    governor_gain_mult=1.90, tau_mult=0.70),
        # Secondary hotspot around bus 37 in the companion paper.
        ClusterRule(hosts=[31, 32, 34, 36, 40, 42], transformer_x_mult=1.00,
                    inertia_mult=1.00, damping_mult=1.00,
                    governor_gain_mult=1.15, tau_mult=0.90),
    ])

    generator_fleet: List[TypeTemplate] = field(default_factory=lambda: [
        TypeTemplate("TB6", 27, 667.0, 6.0, 0.041, 0.033, 0.5),
        TypeTemplate("TC6", 9, 667.0, 8.0, 0.039, 0.033, 0.5),
        TypeTemplate("TB10", 5, 1050.0, 10.0, 0.024, 0.033, 0.5),
        TypeTemplate("TC10", 2, 1050.0, 10.6, 0.021, 0.033, 0.5),
        TypeTemplate("TG3", 11, 300.0, 9.0, 0.065, 0.033, 0.0),
    ])


class PaperLikeCase118V2:
    def __init__(self, cfg: ReproConfigV2):
        self.cfg = cfg
        self.mpc = case118()
        self.bus = self.mpc["bus"].copy()
        self.gen = self.mpc["gen"].copy()
        self.branch = self.mpc["branch"].copy()
        self.host_buses = self.gen[:, 0].astype(int)
        self.n_load = self.bus.shape[0]
        self.n_gen = len(self.host_buses)
        self.n_total = self.n_load + self.n_gen
        self._assign_generator_templates()
        self._build_network()
        self._build_dynamic_model()

    def _assign_generator_templates(self) -> None:
        order = np.argsort(-self.gen[:, 8])
        rows = []
        start = 0
        for templ in self.cfg.generator_fleet:
            subset = order[start:start + templ.count]
            for idx in subset:
                rows.append({
                    "gen_index": int(idx),
                    "host_bus": int(self.host_buses[idx]),
                    "type": templ.name,
                    "capacity_mw": templ.capacity_mw,
                    "capacity_pu": templ.capacity_mw / self.cfg.base_mva,
                    "inertia_s": templ.inertia_s,
                    "xdp": templ.xdp,
                    "deadband_hz": templ.deadband_hz,
                    "delay_s": templ.delay_s,
                })
            start += templ.count
        fleet_df = pd.DataFrame(rows).sort_values("gen_index").reset_index(drop=True)
        self.generator_templates = fleet_df
        self.capacity_pu = fleet_df["capacity_pu"].to_numpy(float)
        self.base_H = fleet_df["inertia_s"].to_numpy(float)
        self.base_xdp = fleet_df["xdp"].to_numpy(float)
        self.type_names = fleet_df["type"].tolist()

    def _build_network(self) -> None:
        B = np.zeros((self.n_total, self.n_total), dtype=float)
        weak_edges = {tuple(sorted(edge)) for edge in self.cfg.north_edges}
        for row in self.branch:
            f = int(row[0])
            t = int(row[1])
            x = float(row[3]) * self.cfg.line_x_scale
            tap = float(row[8]) if row[8] != 0 else 1.0
            b = 1.0 / (x * tap)
            if tuple(sorted((f, t))) in weak_edges:
                b *= self.cfg.north_corridor_scale
            i = f - 1
            j = t - 1
            B[i, i] += b
            B[j, j] += b
            B[i, j] -= b
            B[j, i] -= b

        rule_by_host: Dict[int, ClusterRule] = {}
        for rule in self.cfg.clusters:
            for host in rule.hosts:
                rule_by_host[host] = rule

        tr_x = []
        for k, host in enumerate(self.host_buses):
            x = self.base_xdp[k]
            rule = rule_by_host.get(int(host))
            if rule is not None:
                x *= rule.transformer_x_mult
            tr_x.append(x)
            gi = self.n_load + k
            li = host - 1
            b = 1.0 / x
            B[gi, gi] += b
            B[li, li] += b
            B[gi, li] -= b
            B[li, gi] -= b

        self.transformer_x = np.array(tr_x, dtype=float)
        G = np.arange(self.n_load, self.n_total)
        L = np.arange(self.n_load)
        self.B_full = B
        self.B_GG = B[np.ix_(G, G)]
        self.B_GL = B[np.ix_(G, L)]
        self.B_LG = B[np.ix_(L, G)]
        self.B_LL = B[np.ix_(L, L)]
        self.B_LL_inv = la.inv(self.B_LL)
        self.J = self.B_GG - self.B_GL @ self.B_LL_inv @ self.B_LG
        self.F = -self.B_LL_inv @ self.B_LG
        self.Ldist = self.B_GL @ self.B_LL_inv

    def _build_dynamic_model(self) -> None:
        M = self.cfg.inertia_scale * 2.0 * self.base_H * self.capacity_pu
        D = self.cfg.damping_pu_fraction * self.capacity_pu
        K = self.cfg.governor_gain_scale * self.capacity_pu / self.cfg.droop_R
        Tau = np.full(self.n_gen, self.cfg.governor_tau_s, dtype=float)
        Gamma = np.full(self.n_gen, self.cfg.governor_gamma, dtype=float)

        rule_by_host: Dict[int, ClusterRule] = {}
        for rule in self.cfg.clusters:
            for host in rule.hosts:
                rule_by_host[host] = rule

        for idx, host in enumerate(self.host_buses):
            rule = rule_by_host.get(int(host))
            if rule is None:
                continue
            M[idx] *= rule.inertia_mult
            D[idx] *= rule.damping_mult
            K[idx] *= rule.governor_gain_mult
            Tau[idx] *= rule.tau_mult

        load_for_mu = np.maximum(self.bus[:, 2], self.cfg.min_load_for_mu_mw)
        mu = self.cfg.load_damping_fraction * (load_for_mu / self.cfg.base_mva)

        self.M = M
        self.D = D
        self.K = K
        self.Tau = Tau
        self.Gamma = Gamma
        self.mu = mu
        self.D_tilde = np.diag(D) - self.Ldist @ np.diag(mu) @ self.F
        self.M_inv = np.diag(1.0 / M)
        self.N = np.diag(K * Gamma / M)
        self.A = np.block([
            [np.zeros((self.n_gen, self.n_gen)), self.cfg.base_frequency_hz * np.eye(self.n_gen), np.zeros((self.n_gen, self.n_gen))],
            [-self.M_inv @ self.J, -self.M_inv @ self.D_tilde, self.M_inv],
            [self.N @ self.J, self.N @ self.D_tilde - np.diag(K / Tau), -self.N - np.diag(1.0 / Tau)],
        ])

    def simulate(self) -> Dict[str, np.ndarray]:
        dist_idx = self.cfg.disturbance_load_bus - 1
        disturbance_pu = self.cfg.disturbance_mw / self.cfg.base_mva
        u = self.Ldist[:, dist_idx] * disturbance_pu
        bstep = np.concatenate([np.zeros(self.n_gen), self.M_inv @ u, -self.N @ u])

        eigvals, V = la.eig(self.A)
        V_inv = la.inv(V)
        coeff = V_inv @ bstep
        Cg = np.hstack([np.zeros((self.n_gen, self.n_gen)), np.eye(self.n_gen), np.zeros((self.n_gen, self.n_gen))])
        Cl = self.F @ Cg
        Xg = Cg @ V
        Xl = Cl @ V

        t = np.linspace(0.0, self.cfg.duration_s, self.cfg.n_points)
        modal_step = np.empty((len(eigvals), len(t)), dtype=complex)
        for k, lam in enumerate(eigvals):
            if abs(lam) < 1e-8:
                modal_step[k, :] = t
            else:
                modal_step[k, :] = (np.exp(lam * t) - 1.0) / lam

        resp_g = (Xg @ (modal_step * coeff[:, None])).real
        resp_l = (Xl @ (modal_step * coeff[:, None])).real
        coi = (self.M[:, None] * resp_g).sum(axis=0) / self.M.sum()

        i1 = np.where((np.abs(eigvals) >= 1e-8) & (np.abs(eigvals.imag) < 1e-8))[0]
        i2 = np.where(np.abs(eigvals.imag) >= 1e-8)[0]
        resp_g_global = (Xg[:, i1] @ (modal_step[i1, :] * coeff[i1, None])).real if len(i1) else np.zeros_like(resp_g)
        resp_g_local = (Xg[:, i2] @ (modal_step[i2, :] * coeff[i2, None])).real if len(i2) else np.zeros_like(resp_g)
        coi_global = (self.M[:, None] * resp_g_global).sum(axis=0) / self.M.sum()
        coi_local = (self.M[:, None] * resp_g_local).sum(axis=0) / self.M.sum()
        amp_local = (Xl[:, i2] * coeff[i2]) / eigvals[i2] if len(i2) else np.zeros((self.n_load, 1), dtype=complex)
        indicator = np.linalg.norm(amp_local, axis=1).real

        return {
            "t": t,
            "eigvals": eigvals,
            "resp_gen_pu": resp_g,
            "resp_load_pu": resp_l,
            "coi_pu": coi,
            "coi_global_pu": coi_global,
            "coi_local_pu": coi_local,
            "rocof0_load_hz_per_s": (self.F @ (self.M_inv @ u)) * self.cfg.base_frequency_hz,
            "nadir_load_hz": resp_l.min(axis=1) * self.cfg.base_frequency_hz,
            "indicator": indicator,
        }

    def metrics_table(self, result: Dict[str, np.ndarray]) -> pd.DataFrame:
        df = pd.DataFrame({
            "bus": np.arange(1, self.n_load + 1, dtype=int),
            "nadir_hz": result["nadir_load_hz"],
            "abs_nadir_hz": np.abs(result["nadir_load_hz"]),
            "rocof0_hz_per_s": result["rocof0_load_hz_per_s"],
            "abs_rocof0_hz_per_s": np.abs(result["rocof0_load_hz_per_s"]),
            "indicator": result["indicator"],
        })
        df["nadir_rank"] = df["nadir_hz"].rank(method="dense").astype(int)
        df["rocof_rank"] = df["abs_rocof0_hz_per_s"].rank(method="dense", ascending=False).astype(int)
        df["indicator_rank"] = df["indicator"].rank(method="dense", ascending=False).astype(int)
        return df.sort_values("bus").reset_index(drop=True)

    def modified_branch_table(self) -> pd.DataFrame:
        rows = []
        weak_edges = {tuple(sorted(edge)) for edge in self.cfg.north_edges}
        for idx, row in enumerate(self.branch, start=1):
            f = int(row[0]); t = int(row[1])
            x_old = float(row[3])
            scale = self.cfg.line_x_scale
            if tuple(sorted((f, t))) in weak_edges:
                scale *= 1.0 / self.cfg.north_corridor_scale
            x_new = x_old * self.cfg.line_x_scale
            rows.append({
                "branch_index": idx,
                "from_bus": f,
                "to_bus": t,
                "x_original": x_old,
                "x_modified": x_new,
                "north_corridor_edge": tuple(sorted((f, t))) in weak_edges,
            })
        return pd.DataFrame(rows)

    def generator_table(self) -> pd.DataFrame:
        df = self.generator_templates.copy()
        df["transformer_x_used"] = self.transformer_x
        df["M_used"] = self.M
        df["D_used"] = self.D
        df["K_used"] = self.K
        df["Tau_used_s"] = self.Tau
        df["Gamma_used"] = self.Gamma
        return df


# ----- reporting helpers -----

def save_figures(model: PaperLikeCase118V2, result: Dict[str, np.ndarray], out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    t = result["t"]
    hz = model.cfg.base_frequency_hz
    coi = result["coi_pu"] * hz
    coi_global = result["coi_global_pu"] * hz
    coi_local = result["coi_local_pu"] * hz
    load_hz = result["resp_load_pu"] * hz
    nadir = result["nadir_load_hz"]
    rocof = result["rocof0_load_hz_per_s"]
    indicator = result["indicator"]
    buses = np.arange(1, model.n_load + 1)

    fig, ax1 = plt.subplots(figsize=(8.5, 4.8))
    ax1.plot(t, coi_global, label="global component")
    ax1.plot(t, coi, linestyle="--", label="COI")
    ax1.set_xlabel("Time / s")
    ax1.set_ylabel("Global / COI component, Hz")
    ax2 = ax1.twinx()
    ax2.plot(t, coi_local, label="local component")
    ax2.set_ylabel("Local component, Hz")
    h1, l1 = ax1.get_legend_handles_labels()
    h2, l2 = ax2.get_legend_handles_labels()
    ax1.legend(h1 + h2, l1 + l2, loc="best")
    ax1.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_dir / "fig01_coi_decomposition.png", dpi=200)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(8.5, 4.8))
    selected = [9, 8, 48, 49]
    styles = {9: "-", 8: "-", 48: "--", 49: ":"}
    for bus in selected:
        ax.plot(t, load_hz[bus - 1], linestyle=styles[bus], label=f"Bus {bus:02d}")
    ax.plot(t, coi, linewidth=2.0, label="COI")
    ax.set_xlabel("Time / s")
    ax.set_ylabel("Frequency deviation / Hz")
    ax.grid(True, alpha=0.3)
    ax.legend(loc="best")
    fig.tight_layout()
    fig.savefig(out_dir / "fig02_selected_buses.png", dpi=200)
    plt.close(fig)

    fig, ax1 = plt.subplots(figsize=(10.0, 4.8))
    ax1.bar(buses, np.abs(rocof), width=0.8)
    ax1.set_xlabel("Load bus")
    ax1.set_ylabel("|RoCoF(0+)| / Hz/s")
    ax2 = ax1.twinx()
    ax2.plot(buses, np.abs(nadir), linewidth=1.6)
    ax2.set_ylabel("|Frequency nadir| / Hz")
    fig.tight_layout()
    fig.savefig(out_dir / "fig03_rocof_vs_nadir.png", dpi=200)
    plt.close(fig)

    fig, ax1 = plt.subplots(figsize=(10.0, 4.8))
    ax1.bar(buses, indicator, width=0.8)
    ax1.set_xlabel("Load bus")
    ax1.set_ylabel("Indicator s_ij")
    ax2 = ax1.twinx()
    ax2.plot(buses, np.abs(nadir), linewidth=1.6)
    ax2.set_ylabel("|Frequency nadir| / Hz")
    fig.tight_layout()
    fig.savefig(out_dir / "fig04_indicator_vs_nadir.png", dpi=200)
    plt.close(fig)


def build_summary(model: PaperLikeCase118V2, result: Dict[str, np.ndarray], metrics: pd.DataFrame) -> Dict[str, object]:
    coi = result["coi_pu"] * model.cfg.base_frequency_hz
    coi_global = result["coi_global_pu"] * model.cfg.base_frequency_hz
    coi_local = result["coi_local_pu"] * model.cfg.base_frequency_hz
    return {
        "config": asdict(model.cfg),
        "system": {
            "load_buses": int(model.n_load),
            "generator_side_buses": int(model.n_gen),
            "state_order": int(3 * model.n_gen),
            "base_mva": float(model.cfg.base_mva),
        },
        "achieved_metrics": {
            "coi_nadir_hz": float(coi.min()),
            "coi_final_hz": float(coi[-1]),
            "coi_local_to_global_peak_ratio": float(np.max(np.abs(coi_local)) / max(np.max(np.abs(coi_global)), 1e-12)),
            "bus08_nadir_hz": float(metrics.loc[metrics.bus == 8, "nadir_hz"].iloc[0]),
            "bus09_nadir_hz": float(metrics.loc[metrics.bus == 9, "nadir_hz"].iloc[0]),
            "bus48_nadir_hz": float(metrics.loc[metrics.bus == 48, "nadir_hz"].iloc[0]),
            "bus49_nadir_hz": float(metrics.loc[metrics.bus == 49, "nadir_hz"].iloc[0]),
            "bus08_rocof0_hz_per_s": float(metrics.loc[metrics.bus == 8, "rocof0_hz_per_s"].iloc[0]),
            "bus09_rocof0_hz_per_s": float(metrics.loc[metrics.bus == 9, "rocof0_hz_per_s"].iloc[0]),
            "bus48_rocof0_hz_per_s": float(metrics.loc[metrics.bus == 48, "rocof0_hz_per_s"].iloc[0]),
            "bus49_rocof0_hz_per_s": float(metrics.loc[metrics.bus == 49, "rocof0_hz_per_s"].iloc[0]),
        },
        "top10_abs_rocof_buses": metrics.nlargest(10, "abs_rocof0_hz_per_s")[["bus", "rocof0_hz_per_s"]].to_dict(orient="records"),
        "top10_lowest_nadir_buses": metrics.nsmallest(10, "nadir_hz")[["bus", "nadir_hz"]].to_dict(orient="records"),
        "top10_indicator_buses": metrics.nlargest(10, "indicator")[["bus", "indicator"]].to_dict(orient="records"),
    }


def write_excel(model: PaperLikeCase118V2, metrics: pd.DataFrame, summary: Dict[str, object], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    input_fill = PatternFill("solid", fgColor="E2F0D9")
    thin_gray = Side(style="thin", color="B7B7B7")
    border = Border(bottom=thin_gray)

    ws["A1"] = "Modified IEEE-118 nodal-frequency reconstruction"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws.merge_cells("A1:D1")
    ws["A3"] = "Key inputs"
    ws["A3"].font = Font(bold=True)
    ws["A3"].fill = section_fill
    inputs = [
        ("Base MVA", model.cfg.base_mva),
        ("Base frequency (Hz)", model.cfg.base_frequency_hz),
        ("Line x scale", model.cfg.line_x_scale),
        ("Disturbance bus", model.cfg.disturbance_load_bus),
        ("Disturbance (MW)", model.cfg.disturbance_mw),
    ]
    for r, (k, v) in enumerate(inputs, start=4):
        ws[f"A{r}"] = k
        ws[f"B{r}"] = v
        ws[f"B{r}"].fill = input_fill

    ws["A10"] = "Key outputs"
    ws["A10"].font = Font(bold=True)
    ws["A10"].fill = section_fill
    outputs = [
        ("COI nadir (Hz)", summary["achieved_metrics"]["coi_nadir_hz"]),
        ("COI final @20s (Hz)", summary["achieved_metrics"]["coi_final_hz"]),
        ("Bus 48 RoCoF (Hz/s)", summary["achieved_metrics"]["bus48_rocof0_hz_per_s"]),
        ("Bus 49 RoCoF (Hz/s)", summary["achieved_metrics"]["bus49_rocof0_hz_per_s"]),
        ("Bus 08 nadir (Hz)", summary["achieved_metrics"]["bus08_nadir_hz"]),
        ("Bus 09 nadir (Hz)", summary["achieved_metrics"]["bus09_nadir_hz"]),
    ]
    for r, (k, v) in enumerate(outputs, start=11):
        ws[f"A{r}"] = k
        ws[f"B{r}"] = v

    ws["A19"] = "Checks"
    ws["A19"].font = Font(bold=True)
    ws["A19"].fill = section_fill
    ws["A20"] = "Max |RoCoF|"
    ws["B20"] = "=MAX(Load_Metrics!E2:E119)"
    ws["A21"] = "Worst nadir"
    ws["B21"] = "=MIN(Load_Metrics!B2:B119)"
    ws["A22"] = "Top RoCoF bus"
    ws["B22"] = "=INDEX(Load_Metrics!A2:A119,MATCH(MAX(Load_Metrics!E2:E119),Load_Metrics!E2:E119,0))"

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 26 if col == 1 else 18
    for row in range(1, 23):
        ws[f"A{row}"].alignment = Alignment(vertical="center")
        ws[f"B{row}"].alignment = Alignment(vertical="center")
        if row not in (1, 3, 10, 19):
            ws[f"A{row}"].border = border
            ws[f"B{row}"].border = border

    # detail sheets
    detail_map = {
        "Load_Metrics": metrics,
        "Buses_118": pd.DataFrame(model.bus, columns=[
            "BUS_I", "BUS_TYPE", "PD", "QD", "GS", "BS", "BUS_AREA",
            "VM", "VA", "BASE_KV", "ZONE", "VMAX", "VMIN"
        ]),
        "Branches_Modified": model.modified_branch_table(),
        "Generators_54": model.generator_table(),
    }
    for name, df in detail_map.items():
        ws2 = wb.create_sheet(title=name)
        for c, col in enumerate(df.columns, start=1):
            cell = ws2.cell(row=1, column=c, value=str(col))
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for r, row in enumerate(df.itertuples(index=False), start=2):
            for c, val in enumerate(row, start=1):
                ws2.cell(row=r, column=c, value=float(val) if isinstance(val, (np.floating, np.integer)) else val)
        for i, col in enumerate(df.columns, start=1):
            max_len = max(len(str(col)), *(len(str(x)) for x in df[col].head(40).tolist()))
            ws2.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 12), 24)
        ws2.freeze_panes = "A2"

    wb.save(out_path)


def write_readme(out_dir: Path, summary: Dict[str, object]) -> None:
    text = f"""# Second-pass open-data reconstruction of the TPWRS IEEE-118 case

This run follows the companion Applied Energy paper more closely.

## What changed versus the first pass

- original IEEE-118 branch reactances scaled by 0.4
- 54 generator-side buses use the reported 27xTB6, 9xTC6, 5xTB10, 2xTC10, 11xTG3 fleet mix
- generator-side coupling uses the reported x'd templates as transformer-side coupling priors
- 100 MVA base, 50 Hz conversion and 1% load-damping initialization are explicit
- targeted open-data fitting retained around the north weak area and the 49/50 disturbance corridor

## Achieved headline numbers

- COI nadir: {summary['achieved_metrics']['coi_nadir_hz']:.4f} Hz
- Bus 48 initial RoCoF: {summary['achieved_metrics']['bus48_rocof0_hz_per_s']:.4f} Hz/s
- Bus 49 initial RoCoF: {summary['achieved_metrics']['bus49_rocof0_hz_per_s']:.4f} Hz/s
- Bus 08 nadir: {summary['achieved_metrics']['bus08_nadir_hz']:.4f} Hz
- Bus 09 nadir: {summary['achieved_metrics']['bus09_nadir_hz']:.4f} Hz
- Bus 48 nadir: {summary['achieved_metrics']['bus48_nadir_hz']:.4f} Hz
- Bus 49 nadir: {summary['achieved_metrics']['bus49_nadir_hz']:.4f} Hz
"""
    (out_dir / "README_v2.md").write_text(text, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out_dir", type=Path, default=Path("results_v2"))
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    out_dir = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    model = PaperLikeCase118V2(ReproConfigV2())
    result = model.simulate()
    metrics = model.metrics_table(result)
    metrics.to_csv(out_dir / "load_bus_metrics_v2.csv", index=False)
    model.modified_branch_table().to_csv(out_dir / "branches_modified_v2.csv", index=False)
    model.generator_table().to_csv(out_dir / "generator_dynamic_data_v2.csv", index=False)
    pd.DataFrame(model.bus, columns=[
        "BUS_I", "BUS_TYPE", "PD", "QD", "GS", "BS", "BUS_AREA",
        "VM", "VA", "BASE_KV", "ZONE", "VMAX", "VMIN"
    ]).to_csv(out_dir / "buses_118_original_v2.csv", index=False)
    save_figures(model, result, out_dir)
    summary = build_summary(model, result, metrics)
    with open(out_dir / "summary_v2.json", "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)
    write_excel(model, metrics, summary, out_dir / "modified_case118_v2.xlsx")
    write_readme(out_dir, summary)
    print(json.dumps(summary["achieved_metrics"], indent=2))


if __name__ == "__main__":
    main()
